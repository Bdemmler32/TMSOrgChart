import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── STAFF sheet ──────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Staff"

# Column headers
headers = [
    "id", "name", "title", "ext", "direct", "email", "photo",
    "dept_id", "role", "reports_to_id", "description"
]
header_fill = PatternFill("solid", fgColor="1A2B4A")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Column widths
col_widths = [22, 25, 45, 8, 18, 30, 40, 20, 16, 22, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

# Data rows — role values: executive | deputy | exec_staff | dept_head | manager | staff
# reports_to_id: id of their direct manager (empty = reports to CEO)
rows = [
    # id, name, title, ext, direct, email, photo, dept_id, role, reports_to_id, description
    ("trudi-dunlap","Trudi Dunlap","Executive Director and Chief Executive Officer","275","724-814-3174","tdunlap@tms.org","Dunlap,-Trudi_2025.jpg","executive","executive","",
     "Trudi Dunlap has served as Executive Director and CEO of TMS since 2018, leading the Society's strategic vision, member engagement, and global operations across materials science and engineering."),

    ("adrianne-carolla","Adrianne Carolla","Deputy Executive Director","219","724-814-3180","acarolla@tms.org","Carolla,-Adrianne_2025.jpg","executive","deputy","trudi-dunlap",""),
    ("sarah-huffman","Sarah Huffman","Administrative Assistant","230","724-814-3186","shuffman@tms.org","Huffman,-Sarah_2025.jpg","executive","exec_staff","trudi-dunlap",""),
    ("paul-zappas","Paul Zappas","Senior Manager, Information Technology","223","724-814-3126","zappas@tms.org","Zappas,-Paul_2022.jpg","executive","exec_staff","trudi-dunlap",""),

    # Events dept
    ("diane-scheuring","Diane Scheuring","Department Head, Events, Programming, and Sales","210","724-814-3110","dscheuring@tms.org","Scheuring.-Diane_2025-copy.jpg","events","dept_head","trudi-dunlap",""),
    ("jeff-gnacinski","Jeff Gnacinski","Programming Manager","248","724-814-3166","jgnacinski@tms.org","Gnacinski_Jeffrey_2026.jpg","events","manager","diane-scheuring",""),
    ("tess-killeen","Tess Killeen","Specialty Events Lead","213","724-814-3116","tdejong@tms.org","Killeen,-Tess_2022.jpg","events","manager","diane-scheuring",""),
    ("kellye-parsson","Kellye Parsson","Global Conferences and Events Manager","216","724-814-3194","kparsson@tms.org","Parsson_Kellye_2026.jpg","events","manager","diane-scheuring",""),
    ("colleen-madore","Colleen Madore","Programming and Virtual Events Specialist","227","724-814-3134","cmadore@tms.org","Madore_Colleen_2026.jpg","events","staff","jeff-gnacinski",""),
    ("patricia-warren","Patricia Warren","Programming and Proceedings Specialist","239","724-814-3152","pwarren@tms.org","Warren_Patricia_2026.jpg","events","staff","jeff-gnacinski",""),

    # Member Services dept
    ("courtney-hammer","Courtney Hammer","Department Head, Member Services and Recognition","259","724-814-3172","chammer@tms.org","Hammer,-Courtney_2021.jpg","member-services","dept_head","trudi-dunlap",""),
    ("janel-show","Janel Show","Volunteer and Customer Service Manager","241","724-814-3156","jshow@tms.org","Show_Janel_2026.jpg","member-services","manager","courtney-hammer",""),
    ("deborah-hixon","Deborah Hixon","Awards Program Manager","232","724-814-3142","hixon@tms.org","Hixon_Deborah_2026.jpg","member-services","manager","courtney-hammer",""),
    ("diana-grady","Diana Grady","Customer and Administrative Services Specialist","211","724-814-3112","dgrady@tms.org","Grady,-Diana_2022.jpg","member-services","staff","janel-show",""),
    ("joni-gregg","Joni Gregg","Technical Divisions Administrator","222","724-814-3124","jgregg@tms.org","Gregg,-Joni_2011.jpg","member-services","staff","courtney-hammer",""),

    # Content dept
    ("matt-baker","Matt Baker","Department Head, Content","280","724-814-3176","mbaker@tms.org","Baker,-Matt_2022.jpg","content","dept_head","trudi-dunlap",""),
    ("kelly-markel","Kelly Markel","Publications Managing Editor","281","724-814-3108","kmarkel@tms.org","Markel_Kelly_2026.jpg","content","staff","matt-baker",""),

    # Marketing dept
    ("ashley-anne-bohnert","Ashley-Anne Bohnert","Department Head, Marketing and Communications","224","724-814-3188","abohnert@tms.org","Bohnert_Ashley_2026.jpg","marketing","dept_head","trudi-dunlap",""),
    ("beate-helsel","Beate Helsel","Senior Manager, Research, Engagement, Data, and Information","220","724-814-3182","bhelsel@tms.org","Helsel_Beate_2026.jpg","marketing","manager","ashley-anne-bohnert",""),
    ("kelly-zappas","Kelly Zappas","JOM: The Magazine Editor / Membership Communications Manager","218","724-814-3122","kzappas@tms.org","Zappas,-Kelly_2025.jpg","marketing","manager","ashley-anne-bohnert",""),
    ("megan-enright","Megan Enright","Marketing Administrator","243","724-814-3106","menright@tms.org","Enright,-Megan_2022.jpg","marketing","staff","ashley-anne-bohnert",""),
    ("jillian-schultz","Jillian Schultz","Digital Engagement Specialist","236","724-814-3168","jschultz@tms.org","Schultz_Jillian_2026.jpg","marketing","staff","megan-enright",""),
    ("dave-rasel","Dave Rasel","Senior Manager, Brand and Digital Assets","242","724-814-3158","drasel@tms.org","Rasel,-Dave_2026.jpg","marketing","manager","ashley-anne-bohnert",""),
    ("bob-demmler","Bob Demmler","AI Integration and Visual Communications Specialist","217","724-814-3120","bdemmler@tms.org","Demmler,-Bob_2025.jpg","marketing","staff","dave-rasel",
     "Bob Demmler leads AI integration initiatives and visual communications at TMS, helping the organization leverage emerging technologies to enhance member engagement, content production, and digital storytelling."),
    ("cheryl-geier","Cheryl Geier","Senior Graphic Designer","240","724-814-3154","cgeier@tms.org","Geier,-Cheryl_2022.jpg","marketing","staff","dave-rasel",""),
    ("ken-grzegorczyk","Ken Grzegorczyk","Full Stack Developer and Webmaster","255","724-814-3164","kgrzegorczyk@tms.org","Grzegorczyk,-Ken_2016.jpg","marketing","staff","dave-rasel",""),

    # Finance dept
    ("alicia-arbuckle","Alicia Arbuckle","Controller and Department Head, Finance","235","724-814-3146","aarbuckle@tms.org","Arbuckle,-Alicia_2022.jpg","finance","dept_head","trudi-dunlap",""),
    ("marleen-schrader","Marleen Schrader","Accounting and Grants Manager","245","724-814-3178","mschrader@tms.org","Schrader,-Marleen_2011.jpg","finance","staff","alicia-arbuckle",""),
]

# Row fill alternating
fill_a = PatternFill("solid", fgColor="F8F9FC")
fill_b = PatternFill("solid", fgColor="FFFFFF")
body_font = Font(name="Calibri", size=10)
body_align = Alignment(vertical="top", wrap_text=True)

for r, row in enumerate(rows, 2):
    fill = fill_a if r % 2 == 0 else fill_b
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = fill
        cell.font = body_font
        cell.alignment = body_align
    ws.row_dimensions[r].height = 30

# ── DEPARTMENTS sheet ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Departments")
dept_headers = ["id", "name", "color"]
for col, h in enumerate(dept_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font

dept_rows = [
    ("executive",      "Executive",                      "exec"),
    ("events",         "Events, Programming, and Sales", "yellow"),
    ("member-services","Member Services and Recognition","green"),
    ("content",        "Content",                        "blue-light"),
    ("marketing",      "Marketing and Communications",   "blue"),
    ("finance",        "Finance",                        "pink"),
]
for r, row in enumerate(dept_rows, 2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val).font = body_font

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 15

# ── META sheet ────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Meta")
meta = [
    ("organization", "The Minerals, Metals & Materials Society"),
    ("shortName",    "TMS"),
    ("address",      "5700 Corporate Drive Suite 750, Pittsburgh, PA 15237 USA"),
    ("phone",        "1-724-776-9000"),
    ("fax",          "1-724-776-3770"),
    ("website",      "www.tms.org"),
    ("directoryDate","May 1, 2026"),
    ("version",      "v0.02"),
]
for r, (k, v) in enumerate(meta, 1):
    ws3.cell(row=r, column=1, value=k).font = Font(bold=True, name="Calibri", size=10)
    ws3.cell(row=r, column=2, value=v).font = body_font
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 60

wb.save("/home/claude/tms-org-chart/staff-data.xlsx")
print("Excel saved OK")
